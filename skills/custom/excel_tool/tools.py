"""excel_tool - Excel 文件操作工具"""

import csv
import io
import json
import os
from pathlib import Path
from typing import Any

TOOLS = [
    {
        "name": "read_excel",
        "description": "读取 Excel 文件内容，返回指定 Sheet 的数据表格",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Excel 文件路径（绝对或相对；须英文/ASCII 文件名，勿用中文路径）",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet 名称，不传则读取所有 Sheet",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "最多返回行数，默认全部",
                },
            },
            "required": ["path"],
        },
    },
    {
        "name": "create_excel",
        "description": "创建新的 Excel 文件并写入数据",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "保存路径（绝对或相对；须英文/ASCII 文件名，勿用中文路径）",
                },
                "data": {
                    "type": "string",
                    "description": '写入数据，JSON 字符串格式：{"Sheet1": [["列1", "列2"], ["值1", "值2"]]}，key 为 Sheet 名，value 为行列表',
                },
            },
            "required": ["path", "data"],
        },
    },
    {
        "name": "modify_excel",
        "description": "修改已有 Excel 文件内容（更新单元格、添加 Sheet 或追加行）",
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Excel 文件路径（须英文/ASCII 文件名，勿用中文路径）",
                },
                "action": {
                    "type": "string",
                    "enum": ["update_cell", "append_rows", "add_sheet"],
                    "description": "操作类型：update_cell（更新单元格）、append_rows（追加行）、add_sheet（新增 Sheet）",
                },
                "params": {
                    "type": "string",
                    "description": '操作参数 JSON 字符串。update_cell: {"sheet":"Sheet1","cell":"A1","value":"新值"}；append_rows: {"sheet":"Sheet1","rows":[["a","b"]]}；add_sheet: {"name":"新Sheet","rows":[["列1","列2"]]}',
                },
            },
            "required": ["path", "action", "params"],
        },
    },
]


def _ensure_path(path: str) -> Path:
    """解析路径，支持绝对路径和相对路径（相对于当前工作目录）。"""
    p = Path(path)
    if not p.is_absolute():
        p = Path.cwd() / p
    return p.resolve()


def read_excel(path: str, sheet_name: str | None = None, max_rows: int | None = None) -> str:
    """读取 Excel 文件内容。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "错误：缺少 openpyxl 依赖，请添加 requirements.txt 并安装 openpyxl"

    p = _ensure_path(path)
    if not p.exists():
        return f"错误：文件不存在 — {p}"

    try:
        wb = load_workbook(p, data_only=True, read_only=sheet_name is not None)
    except Exception as e:
        return f"错误：无法打开 Excel 文件 — {e}"

    output_lines = []

    if sheet_name:
        sheets = [sheet_name] if sheet_name in wb.sheetnames else []
        if not sheets:
            wb.close()
            return f"错误：Sheet '{sheet_name}' 不存在，可用 Sheet: {wb.sheetnames}"
    else:
        sheets = wb.sheetnames

    for sn in sheets:
        ws = wb[sn]
        output_lines.append(f"=== Sheet: {sn} ({ws.max_row}行 x {ws.max_column}列) ===")

        # Build header row
        header = []
        for col in range(1, ws.max_column + 1):
            header.append(ws.cell(1, col).value or f"列{col}")
        output_lines.append(" | ".join(str(h) for h in header))
        output_lines.append("-" * max(40, len(header) * 10))

        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_values = [str(v) if v is not None else "" for v in row]
            output_lines.append(" | ".join(row_values))
            row_count += 1
            if max_rows and row_count >= max_rows:
                break

        if max_rows and ws.max_row - 1 > max_rows:
            output_lines.append(f"... 共 {ws.max_row - 1} 行数据，仅显示前 {max_rows} 行")

    wb.close()
    return "\n".join(output_lines)


def create_excel(path: str, data: str) -> str:
    """创建新的 Excel 文件。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "错误：缺少 openpyxl 依赖"

    p = _ensure_path(path)
    if p.exists():
        return f"错误：文件已存在 — {p}，如需覆盖请先删除"

    try:
        parsed = json.loads(data)
        if not isinstance(parsed, dict):
            return "错误：data 须为 JSON 对象，key 为 Sheet 名"
    except json.JSONDecodeError as e:
        return f"错误：data JSON 解析失败 — {e}"

    try:
        wb = Workbook()
        # Remove default sheet if we have custom sheets
        if parsed:
            default_ws = wb.active
            wb.remove(default_ws)

        for sheet_name, rows in parsed.items():
            ws = wb.create_sheet(title=sheet_name)
            for row_idx, row_data in enumerate(rows, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Style header row
                    if row_idx == 1:
                        cell.font = Font(bold=True)

        wb.save(p)
        sheets_list = list(parsed.keys())
        return f"Excel 文件已创建: {p}\nSheet: {sheets_list}"
    except Exception as e:
        return f"错误：创建 Excel 失败 — {e}"


def modify_excel(path: str, action: str, params: str) -> str:
    """修改已有 Excel 文件。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "错误：缺少 openpyxl 依赖"

    p = _ensure_path(path)
    if not p.exists():
        return f"错误：文件不存在 — {p}"

    try:
        kwargs = json.loads(params)
        if not isinstance(kwargs, dict):
            return "错误：params 须为 JSON 对象"
    except json.JSONDecodeError as e:
        return f"错误：params JSON 解析失败 — {e}"

    try:
        wb = load_workbook(p)

        if action == "update_cell":
            sheet_name = kwargs.get("sheet", wb.sheetnames[0])
            cell_ref = kwargs.get("cell", "A1")
            value = kwargs.get("value", "")
            if sheet_name not in wb.sheetnames:
                return f"错误：Sheet '{sheet_name}' 不存在"
            ws = wb[sheet_name]
            ws[cell_ref] = value
            wb.save(p)
            return f"单元格 {cell_ref} 已更新为 '{value}'（Sheet: {sheet_name}）"

        elif action == "append_rows":
            sheet_name = kwargs.get("sheet", wb.sheetnames[0])
            rows = kwargs.get("rows", [])
            if sheet_name not in wb.sheetnames:
                return f"错误：Sheet '{sheet_name}' 不存在"
            ws = wb[sheet_name]
            count = 0
            for row_data in rows:
                ws.append(row_data)
                count += 1
            wb.save(p)
            return f"已追加 {count} 行数据到 Sheet '{sheet_name}'"

        elif action == "add_sheet":
            name = kwargs.get("name", "新Sheet")
            rows = kwargs.get("rows", [])
            if name in wb.sheetnames:
                return f"错误：Sheet '{name}' 已存在"
            ws = wb.create_sheet(title=name)
            for row_data in rows:
                ws.append(row_data)
            wb.save(p)
            return f"已创建新 Sheet '{name}'，写入 {len(rows)} 行"

        else:
            return f"错误：不支持的操作类型 '{action}'"

    except Exception as e:
        return f"错误：修改 Excel 失败 — {e}"
