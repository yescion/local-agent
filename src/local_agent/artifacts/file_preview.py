"""File preview and kind detection (shared by API preview and agent extraction)."""

from __future__ import annotations

import base64
import csv
import io
import mimetypes
from pathlib import Path

_PREVIEW_TEXT_EXTS = {
    ".txt", ".md", ".markdown", ".json", ".yaml", ".yml", ".xml",
    ".html", ".htm", ".css", ".js", ".ts", ".tsx", ".jsx", ".py",
    ".sql", ".log", ".ini", ".toml", ".sh", ".bat", ".ps1",
}
_PREVIEW_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".ico", ".bmp"}
_PREVIEW_SPREADSHEET_EXTS = {".xlsx", ".xlsm", ".xls"}
_PREVIEW_TABLE_EXTS = {".csv", ".tsv"}
_PREVIEW_PDF_EXTS = {".pdf"}
_PREVIEW_DOCX_EXTS = {".docx"}
_PREVIEW_HTML_EXTS = {".html", ".htm"}

_MAX_SPREADSHEET_ROWS = 200
_MAX_SPREADSHEET_COLS = 30
_MAX_SPREADSHEET_SHEETS = 5
_MAX_TABLE_ROWS = 200


def preview_kind(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in _PREVIEW_IMAGE_EXTS:
        return "image"
    if ext in _PREVIEW_SPREADSHEET_EXTS:
        return "spreadsheet"
    if ext in _PREVIEW_TABLE_EXTS:
        return "table"
    if ext in _PREVIEW_PDF_EXTS:
        return "pdf"
    if ext in _PREVIEW_DOCX_EXTS:
        return "docx"
    if ext in _PREVIEW_HTML_EXTS:
        return "html"
    if ext in _PREVIEW_TEXT_EXTS:
        return "text"
    mime, _ = mimetypes.guess_type(str(path))
    if mime and mime.startswith("image/"):
        return "image"
    if mime == "application/pdf":
        return "pdf"
    if mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return "docx"
    if mime in ("text/html", "application/xhtml+xml"):
        return "html"
    if mime and (mime.startswith("text/") or mime in ("application/json", "application/xml")):
        return "text"
    if ext in {".xlsx", ".xls", ".xlsm"}:
        return "spreadsheet"
    return "binary"


def build_spreadsheet_preview(path: Path, *, size_bytes: int) -> dict | None:
    try:
        import openpyxl
    except ImportError:
        return None

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    sheets: list[dict] = []
    truncated = False
    try:
        for sheet_name in wb.sheetnames[:_MAX_SPREADSHEET_SHEETS]:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx >= _MAX_SPREADSHEET_ROWS:
                    truncated = True
                    break
                cells = []
                for col_idx, cell in enumerate(row):
                    if col_idx >= _MAX_SPREADSHEET_COLS:
                        truncated = True
                        break
                    cells.append("" if cell is None else str(cell))
                rows.append(cells)
            sheets.append({"name": sheet_name, "rows": rows})
        if len(wb.sheetnames) > _MAX_SPREADSHEET_SHEETS:
            truncated = True
    finally:
        wb.close()

    if not sheets:
        return None

    return {
        "kind": "spreadsheet",
        "sheets": sheets,
        "truncated": truncated,
        "size_bytes": size_bytes,
    }


def build_table_preview(path: Path, *, size_bytes: int, delimiter: str = ",") -> dict | None:
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return None

    reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
    rows: list[list[str]] = []
    truncated = False
    for row_idx, row in enumerate(reader):
        if row_idx >= _MAX_TABLE_ROWS:
            truncated = True
            break
        rows.append(row[:_MAX_SPREADSHEET_COLS])

    if not rows:
        return None

    return {
        "kind": "table",
        "rows": rows,
        "truncated": truncated,
        "size_bytes": size_bytes,
    }


def build_preview(path: Path, *, max_chars: int = 200_000) -> dict:
    if not path.is_file():
        return {"kind": "missing", "error": "文件不存在"}

    kind = preview_kind(path)
    stat = path.stat()
    mime, _ = mimetypes.guess_type(str(path))
    size_bytes = stat.st_size

    if kind == "image":
        data = base64.b64encode(path.read_bytes()).decode("ascii")
        return {
            "kind": "image",
            "mime": mime or "application/octet-stream",
            "data": data,
            "size_bytes": size_bytes,
        }

    if kind == "spreadsheet":
        preview = build_spreadsheet_preview(path, size_bytes=size_bytes)
        if preview:
            return preview
        return {
            "kind": "binary",
            "mime": mime or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "size_bytes": size_bytes,
            "message": "Excel 预览需要 openpyxl，请安装后重试或下载查看",
        }

    if kind == "table":
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        preview = build_table_preview(path, size_bytes=size_bytes, delimiter=delimiter)
        if preview:
            return preview

    if kind == "pdf":
        return {
            "kind": "pdf",
            "mime": mime or "application/pdf",
            "size_bytes": size_bytes,
            "download_url_hint": True,
        }

    if kind == "docx":
        return {
            "kind": "docx",
            "mime": mime or "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "size_bytes": size_bytes,
            "download_url_hint": True,
        }

    if kind == "html":
        raw = path.read_text(encoding="utf-8", errors="replace")
        truncated = len(raw) > max_chars
        if truncated:
            raw = raw[:max_chars]
        return {
            "kind": "html",
            "mime": mime or "text/html",
            "content": raw,
            "truncated": truncated,
            "size_bytes": size_bytes,
        }

    if kind == "text":
        raw = path.read_text(encoding="utf-8", errors="replace")
        truncated = len(raw) > max_chars
        if truncated:
            raw = raw[:max_chars]
        return {
            "kind": "text",
            "mime": mime or "text/plain",
            "content": raw,
            "truncated": truncated,
            "size_bytes": size_bytes,
        }

    return {
        "kind": "binary",
        "mime": mime or "application/octet-stream",
        "size_bytes": size_bytes,
        "message": "此文件类型不支持内联预览，请下载查看",
    }
